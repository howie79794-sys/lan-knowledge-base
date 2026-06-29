from __future__ import annotations

import csv
import io
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree


class ParseResult:
    def __init__(self, markdown: str, text: str, parser: str) -> None:
        self.markdown = markdown
        self.text = text
        self.parser = parser


def parse_document(file_path: Path, file_format: str) -> ParseResult:
    suffix = file_path.suffix.lower()
    if file_format in {"text", "markdown"}:
        return _parse_plain_text(file_path)
    if file_format == "pdf":
        return _parse_pdf(file_path)
    if file_format == "word" and suffix == ".docx":
        return _parse_docx(file_path)
    if file_format == "ppt" and suffix == ".pptx":
        return _parse_pptx(file_path)
    if file_format == "excel" and suffix == ".xlsx":
        return _parse_xlsx(file_path)
    if file_format == "csv":
        return _parse_csv(file_path)
    raise ValueError(f"暂不支持解析 {suffix} 文件，请先保留原文下载。")


def _parse_plain_text(file_path: Path) -> ParseResult:
    raw = file_path.read_bytes()
    text = raw.decode("utf-8", errors="ignore")
    return ParseResult(markdown=text, text=text, parser="plain-text")


def _parse_pdf(file_path: Path) -> ParseResult:
    try:
        from pypdf import PdfReader
    except Exception as exc:  # pragma: no cover - dependency fallback
        raise ValueError("PDF 解析依赖 pypdf 未安装。") from exc

    reader = PdfReader(str(file_path))
    pages: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        if page_text.strip():
            pages.append(f"## 第 {index} 页\n\n{page_text.strip()}")
    text = "\n\n".join(pages).strip()
    if not text:
        raise ValueError("PDF 未提取到文本，可能是扫描件或图片型 PDF。")
    return ParseResult(markdown=text, text=_plain(text), parser="pypdf")


def _parse_docx(file_path: Path) -> ParseResult:
    with zipfile.ZipFile(file_path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    paragraphs: list[str] = []
    for paragraph in root.iter(_tag("p")):
        parts = [node.text or "" for node in paragraph.iter() if _local_name(node.tag) == "t"]
        line = "".join(parts).strip()
        if line:
            paragraphs.append(line)
    text = "\n\n".join(paragraphs)
    if not text:
        raise ValueError("Word 文档未提取到文本。")
    return ParseResult(markdown=text, text=text, parser="docx-xml")


def _parse_pptx(file_path: Path) -> ParseResult:
    slides: list[str] = []
    with zipfile.ZipFile(file_path) as archive:
        slide_names = sorted(name for name in archive.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml"))
        for index, name in enumerate(slide_names, start=1):
            root = ElementTree.fromstring(archive.read(name))
            parts = [node.text or "" for node in root.iter() if _local_name(node.tag) == "t"]
            text = "\n".join(part.strip() for part in parts if part.strip())
            if text:
                slides.append(f"## 第 {index} 页\n\n{text}")
    markdown = "\n\n".join(slides)
    if not markdown:
        raise ValueError("PPT 未提取到文本。")
    return ParseResult(markdown=markdown, text=_plain(markdown), parser="pptx-xml")


def _parse_xlsx(file_path: Path) -> ParseResult:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - dependency fallback
        raise ValueError("Excel 解析依赖 openpyxl 未安装。") from exc

    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(max_row=80, values_only=True):
            values = ["" if cell is None else str(cell).strip() for cell in row]
            if any(values):
                rows.append(values)
            if len(rows) >= 40:
                break
        if rows:
            sections.append(f"## {sheet.title}\n\n{_markdown_table(rows)}")
    markdown = "\n\n".join(sections)
    if not markdown:
        raise ValueError("Excel 未提取到文本。")
    return ParseResult(markdown=markdown, text=_plain(markdown), parser="openpyxl")


def _parse_csv(file_path: Path) -> ParseResult:
    raw = file_path.read_bytes().decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(raw))
    rows = []
    for row in reader:
        rows.append([cell.strip() for cell in row])
        if len(rows) >= 80:
            break
    if not rows:
        raise ValueError("CSV 未提取到文本。")
    markdown = _markdown_table(rows)
    return ParseResult(markdown=markdown, text=_plain(markdown), parser="csv")


def _markdown_table(rows: list[list[str]]) -> str:
    width = max(len(row) for row in rows)
    normalized = [row + [""] * (width - len(row)) for row in rows]
    header = normalized[0]
    body = normalized[1:] or [[""] * width]
    lines = [
        "| " + " | ".join(_escape_cell(cell) for cell in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(_escape_cell(cell) for cell in row) + " |")
    return "\n".join(lines)


def _escape_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")


def _tag(name: str) -> str:
    return f"{{http://schemas.openxmlformats.org/wordprocessingml/2006/main}}{name}"


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _plain(markdown: str) -> str:
    return re.sub(r"[#*_`|>-]+", " ", markdown).strip()
